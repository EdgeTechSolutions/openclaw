#!/usr/bin/env python3
"""
Create professional Gantt chart Excel for Odin ATVP AI project.
"""
from datetime import date, timedelta
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.views import SheetView, Pane
from openpyxl.worksheet.page import PageMargins

# ─────────────────────────────────────────────────────────────────────────────
# COLOR DEFINITIONS
# ─────────────────────────────────────────────────────────────────────────────
HEADER_BG    = "2D2D2D"
HEADER_FG    = "FFFFFF"
TOTAL_BG     = "2D2D2D"
TOTAL_FG     = "FFFFFF"
GANTT_HEADER_BG = "3D3D3D"

PHASES = {
    "F1": {"dark": "1B4F72", "light": "D4E6F1", "bar": "2980B9"},
    "F2": {"dark": "1A5276", "light": "D6EAF8", "bar": "3498DB"},
    "F3": {"dark": "145A32", "light": "D5F5E3", "bar": "27AE60"},
    "F4": {"dark": "6C3483", "light": "E8DAEF", "bar": "8E44AD"},
    "F5": {"dark": "784212", "light": "FDEBD0", "bar": "E67E22"},
    "F6": {"dark": "1B2631", "light": "D5D8DC", "bar": "566573"},
    "F7": {"dark": "7B241C", "light": "FADBD8", "bar": "E74C3C"},
    "F8": {"dark": "0E6655", "light": "D1F2EB", "bar": "1ABC9C"},
    "F9": {"dark": "1A5276", "light": "AED6F1", "bar": "2E86C1"},
}

# Activity 1.1 spans whole project — use a lighter/more transparent bar color
ACTIVITY_11_BAR = "AED6F1"

# ─────────────────────────────────────────────────────────────────────────────
# GANTT WEEK DEFINITIONS  (26 weeks: 2026-04-01 .. ~2026-10-06)
# ─────────────────────────────────────────────────────────────────────────────
SL_MONTHS = {
    4: "Apr", 5: "Maj", 6: "Jun", 7: "Jul",
    8: "Avg", 9: "Sep", 10: "Okt"
}

def build_weeks():
    """Return list of (label, start_date, end_date) for each week column.
    Week is assigned to the month that contains its midpoint (day 3 = Wednesday)."""
    weeks = []
    month_counter = {}  # month -> week-within-month counter
    d = date(2026, 4, 1)
    while d <= date(2026, 10, 13):
        end = d + timedelta(days=6)
        mid = d + timedelta(days=3)   # midpoint of the week
        m = mid.month
        month_counter[m] = month_counter.get(m, 0) + 1
        label = f"T{month_counter[m]} {SL_MONTHS.get(m, str(m))}"
        weeks.append((label, d, end))
        d += timedelta(days=7)
    return weeks

WEEKS = build_weeks()
NUM_WEEKS = len(WEEKS)
GANTT_START_COL = 8  # column H (1-indexed)

# ─────────────────────────────────────────────────────────────────────────────
# WBS DATA
# ─────────────────────────────────────────────────────────────────────────────
# Each entry: type, phase_key, number, name, description, om, start, end
# type = "phase" or "activity"
# start/end = date objects or None

def d(s):
    return date.fromisoformat(s)

WBS = [
    # Phase F1
    ("phase",    "F1", "F1", "Projektno vodenje in analiza",     "",                                                                 2.00,  None,       None),
    ("activity", "F1", "1.1", "  Projektno vodenje",             "Koordinacija, sestanki, poročanje, tveganja",                     1.0,   d("2026-04-01"), d("2026-09-30")),
    ("activity", "F1", "1.2", "  Podrobna analiza zahtev",       "Delavnice z ATVP, definicija podatkovnih tokov",                  0.5,   d("2026-04-01"), d("2026-04-14")),
    ("activity", "F1", "1.3", "  Arhitekturno načrtovanje",      "Hibridna arhitektura (on-prem + Azure), varnostni model",         0.5,   d("2026-04-15"), d("2026-04-28")),

    # Phase F2
    ("phase",    "F2", "F2", "Infrastruktura in okolje",         "",                                                                 1.50,  None,       None),
    ("activity", "F2", "2.1", "  Vzpostavitev Azure okolja",     "Azure OpenAI, Azure AI Search, Azure AD",                         0.5,   d("2026-04-29"), d("2026-05-12")),
    ("activity", "F2", "2.2", "  On-premise infrastruktura",     "Strežniki, mrežna konfiguracija",                                 0.5,   d("2026-04-29"), d("2026-05-12")),
    ("activity", "F2", "2.3", "  CI/CD in razvojna okolja",      "Deployment pipeline, testna okolja",                              0.5,   d("2026-05-13"), d("2026-05-26")),

    # Phase F3
    ("phase",    "F3", "F3", "Backend razvoj",                   "",                                                                 2.50,  None,       None),
    ("activity", "F3", "3.1", "  API ogrodje (Django)",          "Django, avtentikacija, avtorizacija, revizijska sled",             0.5,   d("2026-05-27"), d("2026-06-09")),
    ("activity", "F3", "3.2", "  Modul za delovne skupine",      "CRUD, OneDrive sinhronizacija",                                   0.5,   d("2026-06-10"), d("2026-06-23")),
    ("activity", "F3", "3.3", "  Modul za transkripte",          "Uvoz videov, SRC integracija, urejanje",                          0.5,   d("2026-06-10"), d("2026-06-23")),
    ("activity", "F3", "3.4", "  RAG sistem (Qdrant)",           "Vektorizacija, indeksiranje, zakonodajna baza",                    0.5,   d("2026-06-24"), d("2026-07-07")),
    ("activity", "F3", "3.5", "  Konfiguracija agentov",         "Prompt, dokumenti, orodja, LLM model",                            0.5,   d("2026-07-08"), d("2026-07-21")),

    # Phase F4
    ("phase",    "F4", "F4", "AI asistenti",                     "",                                                                 1.75,  None,       None),
    ("activity", "F4", "4.1", "  Asistent za povzemanje",        "Izbira dokumentov, generiranje povzetkov",                        0.5,   d("2026-07-22"), d("2026-08-04")),
    ("activity", "F4", "4.2", "  Asistent za ref. zakonodaje",   "Označevanje, barvno kodiranje, časovna relevantnost",              0.75,  d("2026-08-05"), d("2026-08-25")),
    ("activity", "F4", "4.3", "  Asistent za ref. transkripta",  "Povezovanje izjav z dokumentacijo",                               0.5,   d("2026-08-05"), d("2026-08-18")),

    # Phase F5
    ("phase",    "F5", "F5", "Frontend razvoj",                  "",                                                                 1.375, None,       None),
    ("activity", "F5", "5.1", "  Ogrodje in navigacija",         "React, Azure AD prijava, routing",                                0.25,  d("2026-06-10"), d("2026-06-16")),
    ("activity", "F5", "5.2", "  Nadzorna plošča",               "Kartice delovnih skupin, obvestila",                              0.125, d("2026-06-17"), d("2026-06-23")),
    ("activity", "F5", "5.3", "  Delovna skupina — vmesniki",    "Dokumenti, člani, transkripti, asistenti",                        0.125, d("2026-06-24"), d("2026-06-30")),
    ("activity", "F5", "5.4", "  Pregled transkriptov",          "Govorci, časovne oznake, inline urejanje",                        0.25,  d("2026-07-01"), d("2026-07-07")),
    ("activity", "F5", "5.5", "  Vmesniki asistentov",           "Split-pane, barvno označevanje",                                  0.25,  d("2026-08-05"), d("2026-08-11")),
    ("activity", "F5", "5.6", "  Klepetalnik (Chatbot)",         "Floating widget, kontekstno odgovarjanje",                        0.25,  d("2026-08-12"), d("2026-08-18")),
    ("activity", "F5", "5.7", "  Administracija",                "Uporabniki, zakonodaja, revizijska sled, agenti",                  0.125, d("2026-07-01"), d("2026-07-07")),

    # Phase F6
    ("phase",    "F6", "F6", "Integracije",                      "",                                                                 0.625, None,       None),
    ("activity", "F6", "6.1", "  Azure AD (SSO)",                "Avtentikacija in avtorizacija",                                   0.25,  d("2026-06-10"), d("2026-06-16")),
    ("activity", "F6", "6.2", "  OneDrive sinhronizacija",       "Dvosmerna sinhronizacija",                                        0.125, d("2026-06-24"), d("2026-06-30")),
    ("activity", "F6", "6.3", "  SRC transkripcija",             "Govor v besedilo integracija",                                    0.25,  d("2026-06-24"), d("2026-07-07")),

    # Phase F7
    ("phase",    "F7", "F7", "Skladnost in varnost",             "",                                                                 0.375, None,       None),
    ("activity", "F7", "7.1", "  GDPR skladnost",                "Klasifikacija podatkov, DPA",                                     0.125, d("2026-04-15"), d("2026-04-21")),
    ("activity", "F7", "7.2", "  EU AI Act skladnost",           "Dokumentiranje tveganj, razložljivost",                           0.25,  d("2026-04-22"), d("2026-05-05")),

    # Phase F8
    ("phase",    "F8", "F8", "Testiranje",                       "",                                                                 1.00,  None,       None),
    ("activity", "F8", "8.1", "  Funkcionalno testiranje",       "Testiranje vseh modulov",                                         0.25,  d("2026-08-26"), d("2026-09-01")),
    ("activity", "F8", "8.2", "  AI testiranje",                 "Evalvacija kakovosti AI",                                         0.25,  d("2026-09-02"), d("2026-09-08")),
    ("activity", "F8", "8.4", "  UAT (uporabniško testiranje)",  "Testiranje z ATVP uporabniki",                                    0.5,   d("2026-09-09"), d("2026-09-22")),

    # Phase F9
    ("phase",    "F9", "F9", "Uvedba in izobraževanje",          "",                                                                 1.00,  None,       None),
    ("activity", "F9", "9.1", "  Uvedba v produkcijo",           "Namestitev, migracija",                                           0.25,  d("2026-09-23"), d("2026-09-29")),
    ("activity", "F9", "9.2", "  Izobraževanje uporabnikov",     "Delavnice za ATVP zaposlene",                                     0.5,   d("2026-09-30"), d("2026-10-13")),
    ("activity", "F9", "9.3", "  Dokumentacija",                 "Tehnična dokumentacija, navodila",                                0.25,  d("2026-09-09"), d("2026-09-15")),
]

# ─────────────────────────────────────────────────────────────────────────────
# HELPER FUNCTIONS
# ─────────────────────────────────────────────────────────────────────────────
def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def bold_font(size=10, color="000000"):
    return Font(bold=True, size=size, color=color, name="Calibri")

def regular_font(size=10, color="000000"):
    return Font(bold=False, size=size, color=color, name="Calibri")

THIN = Side(border_style="thin", color="BDBDBD")
THICK = Side(border_style="medium", color="888888")

def thin_border():
    return Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def activity_weeks_active(act_start, act_end, week_start, week_end):
    """Return True if activity overlaps with the given week."""
    return act_start <= week_end and act_end >= week_start

def duration_weeks(start, end):
    return round((end - start).days / 7 + 1, 1)

# ─────────────────────────────────────────────────────────────────────────────
# BUILD WORKBOOK
# ─────────────────────────────────────────────────────────────────────────────
wb = Workbook()
ws = wb.active
ws.title = "Odin – Gantt"

# ── Page setup ────────────────────────────────────────────────────────────────
ws.page_setup.orientation = "landscape"
ws.page_setup.fitToPage = True
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0
ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)

# ── Title row (row 1) ──────────────────────────────────────────────────────
ws.row_dimensions[1].height = 24
ws.merge_cells(f"A1:{get_column_letter(GANTT_START_COL + NUM_WEEKS - 1)}1")
title_cell = ws["A1"]
title_cell.value = "PROJEKT ODIN – Agencija za trg vrednostnih papirjev (ATVP)  |  Načrt izvajanja 2026"
title_cell.font = Font(bold=True, size=14, color="FFFFFF", name="Calibri")
title_cell.fill = fill("1B2631")
title_cell.alignment = Alignment(horizontal="center", vertical="center")

# ── Header row (row 2) ──────────────────────────────────────────────────────
HEADER_ROW = 2
ws.row_dimensions[HEADER_ROW].height = 28

headers = ["#", "Faza / Aktivnost", "Opis", "Oseba-meseci", "Začetek", "Konec", "Trajanje (tedni)"]
for ci, h in enumerate(headers, start=1):
    cell = ws.cell(row=HEADER_ROW, column=ci, value=h)
    cell.fill = fill(HEADER_BG)
    cell.font = bold_font(10, HEADER_FG)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border()

# Gantt week headers
for wi, (label, wstart, wend) in enumerate(WEEKS):
    ci = GANTT_START_COL + wi
    cell = ws.cell(row=HEADER_ROW, column=ci, value=label)
    cell.fill = fill(GANTT_HEADER_BG)
    cell.font = Font(bold=True, size=7, color="FFFFFF", name="Calibri")
    cell.alignment = Alignment(horizontal="center", vertical="center", text_rotation=90)
    cell.border = thin_border()

# ── Data rows ────────────────────────────────────────────────────────────────
DATA_START_ROW = 3

for row_idx, entry in enumerate(WBS):
    row_type, phase_key, num, name, desc, om, act_start, act_end = entry
    excel_row = DATA_START_ROW + row_idx
    ws.row_dimensions[excel_row].height = 16

    p = PHASES[phase_key]

    if row_type == "phase":
        bg = p["dark"]
        fg = "FFFFFF"
        fnt = bold_font(10, fg)
        om_display = f"{om:.2f} OM"
    else:
        bg = p["light"]
        fg = "000000"
        fnt = regular_font(10, fg)
        om_display = f"{om:.3f}".rstrip("0").rstrip(".")

    # Column A: number
    c = ws.cell(row=excel_row, column=1, value=num)
    c.fill = fill(bg); c.font = fnt; c.border = thin_border()
    c.alignment = Alignment(horizontal="center", vertical="center")

    # Column B: name
    c = ws.cell(row=excel_row, column=2, value=name)
    c.fill = fill(bg); c.font = fnt; c.border = thin_border()
    c.alignment = Alignment(horizontal="left", vertical="center", indent=0)

    # Column C: description
    c = ws.cell(row=excel_row, column=3, value=desc)
    c.fill = fill(bg); c.font = regular_font(9, fg); c.border = thin_border()
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Column D: OM
    c = ws.cell(row=excel_row, column=4, value=om_display)
    c.fill = fill(bg); c.font = fnt; c.border = thin_border()
    c.alignment = Alignment(horizontal="center", vertical="center")

    # Column E: start date
    if act_start:
        c = ws.cell(row=excel_row, column=5, value=act_start.strftime("%d.%m.%Y"))
    else:
        c = ws.cell(row=excel_row, column=5, value="")
    c.fill = fill(bg); c.font = regular_font(9, fg); c.border = thin_border()
    c.alignment = Alignment(horizontal="center", vertical="center")

    # Column F: end date
    if act_end:
        c = ws.cell(row=excel_row, column=6, value=act_end.strftime("%d.%m.%Y"))
    else:
        c = ws.cell(row=excel_row, column=6, value="")
    c.fill = fill(bg); c.font = regular_font(9, fg); c.border = thin_border()
    c.alignment = Alignment(horizontal="center", vertical="center")

    # Column G: duration in weeks
    if act_start and act_end:
        dur = duration_weeks(act_start, act_end)
        c = ws.cell(row=excel_row, column=7, value=dur)
    else:
        c = ws.cell(row=excel_row, column=7, value="")
    c.fill = fill(bg); c.font = regular_font(9, fg); c.border = thin_border()
    c.alignment = Alignment(horizontal="center", vertical="center")

    # ── Gantt columns ────────────────────────────────────────────────────────
    for wi, (label, wstart, wend) in enumerate(WEEKS):
        ci = GANTT_START_COL + wi
        cell = ws.cell(row=excel_row, column=ci, value="")
        cell.border = thin_border()

        if row_type == "phase":
            # Phase rows: empty / same dark bg
            cell.fill = fill(bg)
        else:
            # Activity row
            if act_start and act_end:
                is_active = activity_weeks_active(act_start, act_end, wstart, wend)
                if is_active:
                    # Special case: activity 1.1 uses lighter bar color
                    if num.strip() == "1.1":
                        bar_color = ACTIVITY_11_BAR
                    else:
                        bar_color = p["bar"]
                    cell.fill = fill(bar_color)
                else:
                    cell.fill = fill("F5F5F5")
            else:
                cell.fill = fill("F5F5F5")

# ── Total row ─────────────────────────────────────────────────────────────────
total_row = DATA_START_ROW + len(WBS)
ws.row_dimensions[total_row].height = 18

total_om = 12.125
labels = ["", "SKUPAJ / TOTAL", "", f"{total_om:.3f} OM", "", "", ""]
for ci, val in enumerate(labels, start=1):
    c = ws.cell(row=total_row, column=ci, value=val)
    c.fill = fill(TOTAL_BG)
    c.font = bold_font(10, TOTAL_FG)
    c.border = thin_border()
    c.alignment = Alignment(horizontal="center" if ci != 2 else "left", vertical="center")

for wi in range(NUM_WEEKS):
    ci = GANTT_START_COL + wi
    c = ws.cell(row=total_row, column=ci, value="")
    c.fill = fill(TOTAL_BG)
    c.border = thin_border()

# ─────────────────────────────────────────────────────────────────────────────
# COLUMN WIDTHS
# ─────────────────────────────────────────────────────────────────────────────
ws.column_dimensions["A"].width = 8
ws.column_dimensions["B"].width = 36
ws.column_dimensions["C"].width = 46
ws.column_dimensions["D"].width = 14
ws.column_dimensions["E"].width = 12
ws.column_dimensions["F"].width = 12
ws.column_dimensions["G"].width = 14

for wi in range(NUM_WEEKS):
    col_letter = get_column_letter(GANTT_START_COL + wi)
    ws.column_dimensions[col_letter].width = 4.2

# ─────────────────────────────────────────────────────────────────────────────
# FREEZE PANES  (freeze row 1+2 and column A)
# ─────────────────────────────────────────────────────────────────────────────
ws.freeze_panes = "B3"

# ─────────────────────────────────────────────────────────────────────────────
# LEGEND block (below total row)
# ─────────────────────────────────────────────────────────────────────────────
legend_row = total_row + 2
ws.cell(row=legend_row, column=1, value="Legenda:").font = bold_font(9, "2D2D2D")
ws.cell(row=legend_row, column=1).alignment = Alignment(vertical="center")

phase_labels = [
    ("F1", "Projektno vodenje in analiza"),
    ("F2", "Infrastruktura in okolje"),
    ("F3", "Backend razvoj"),
    ("F4", "AI asistenti"),
    ("F5", "Frontend razvoj"),
    ("F6", "Integracije"),
    ("F7", "Skladnost in varnost"),
    ("F8", "Testiranje"),
    ("F9", "Uvedba in izobraževanje"),
]

for i, (pk, plabel) in enumerate(phase_labels):
    lr = legend_row + 1 + i
    ws.row_dimensions[lr].height = 14
    swatch = ws.cell(row=lr, column=1, value="  " + pk)
    swatch.fill = fill(PHASES[pk]["bar"])
    swatch.font = bold_font(8, "FFFFFF")
    swatch.alignment = Alignment(horizontal="center", vertical="center")

    lbl = ws.cell(row=lr, column=2, value=plabel)
    lbl.font = regular_font(8)
    lbl.alignment = Alignment(vertical="center")

# ─────────────────────────────────────────────────────────────────────────────
# SAVE
# ─────────────────────────────────────────────────────────────────────────────
import os
os.makedirs("/workspace/shared", exist_ok=True)
output_path = "/workspace/shared/odin-gantt.xlsx"
wb.save(output_path)
print(f"Saved: {output_path}")
print(f"Weeks: {NUM_WEEKS}")
print(f"Data rows: {len(WBS)}")
