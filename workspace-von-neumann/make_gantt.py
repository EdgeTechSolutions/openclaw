#!/usr/bin/env python3
"""
Odin Project – Professional Excel Gantt Chart Generator
"""

from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.views import SheetView, Pane

# ─────────────────────────────────────────────
# Colour helpers
# ─────────────────────────────────────────────
def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color.lstrip("#"))

def thin_border():
    s = Side(style="thin", color="BBBBBB")
    return Border(left=s, right=s, top=s, bottom=s)

def medium_border():
    s = Side(style="medium", color="888888")
    return Border(left=s, right=s, top=s, bottom=s)

# ─────────────────────────────────────────────
# Phase definitions
# ─────────────────────────────────────────────
PHASES = {
    "F1": {"phase_bg": "1B4F72", "act_bg": "D4E6F1", "bar": "2980B9", "bar_light": "AED6F1"},
    "F2": {"phase_bg": "1A5276", "act_bg": "D6EAF8", "bar": "3498DB"},
    "F3": {"phase_bg": "145A32", "act_bg": "D5F5E3", "bar": "27AE60"},
    "F4": {"phase_bg": "6C3483", "act_bg": "E8DAEF", "bar": "8E44AD"},
    "F5": {"phase_bg": "784212", "act_bg": "FDEBD0", "bar": "E67E22"},
    "F6": {"phase_bg": "1B2631", "act_bg": "D5D8DC", "bar": "566573"},
    "F7": {"phase_bg": "7B241C", "act_bg": "FADBD8", "bar": "E74C3C"},
    "F8": {"phase_bg": "0E6655", "act_bg": "D1F2EB", "bar": "1ABC9C"},
    "F9": {"phase_bg": "1A5276", "act_bg": "AED6F1", "bar": "2E86C1"},
}

# ─────────────────────────────────────────────
# WBS data
# Each entry: (row_type, phase_key, number, name, description, pm, t_start, t_end)
# row_type: 'phase' | 'activity'
# t_start / t_end: integers (1-26), or None
# ─────────────────────────────────────────────
WBS = [
    # F1
    ("phase",    "F1", "F1", "Projektno vodenje in analiza",     "",                                               2.000, None, None),
    ("activity", "F1", "1.1","Projektno vodenje",                "Koordinacija, sestanki, poročanje",             1.000, 1,   26),
    ("activity", "F1", "1.2","Podrobna analiza zahtev",          "Delavnice z ATVP, podatkovni tokovi",           0.500, 1,    2),
    ("activity", "F1", "1.3","Arhitekturno načrtovanje",         "Hibridna arhitektura, varnostni model",         0.500, 3,    4),
    # F2
    ("phase",    "F2", "F2", "Infrastruktura in okolje",         "",                                               1.500, None, None),
    ("activity", "F2", "2.1","Vzpostavitev Azure okolja",        "Azure OpenAI, AI Search, AD",                   0.500, 5,    6),
    ("activity", "F2", "2.2","On-premise infrastruktura",        "Strežniki, mrežna konfiguracija",               0.500, 5,    6),
    ("activity", "F2", "2.3","CI/CD in razvojna okolja",         "Pipeline, testna okolja",                       0.500, 7,    8),
    # F3
    ("phase",    "F3", "F3", "Backend razvoj",                   "",                                               2.500, None, None),
    ("activity", "F3", "3.1","API ogrodje (Django)",             "Avtentikacija, avtorizacija, revizijska sled",  0.500, 9,   10),
    ("activity", "F3", "3.2","Modul za delovne skupine",         "CRUD, OneDrive sinhronizacija",                 0.500, 11,  12),
    ("activity", "F3", "3.3","Modul za transkripte",             "Uvoz videov, SRC, urejanje",                    0.500, 11,  12),
    ("activity", "F3", "3.4","RAG sistem (Qdrant)",              "Vektorizacija, indeksiranje",                   0.500, 13,  14),
    ("activity", "F3", "3.5","Konfiguracija agentov",            "Prompt, dokumenti, orodja, LLM",               0.500, 15,  16),
    # F4
    ("phase",    "F4", "F4", "AI asistenti",                     "",                                               1.750, None, None),
    ("activity", "F4", "4.1","Asistent za povzemanje",           "Izbira dokumentov, generiranje povzetkov",      0.500, 17,  18),
    ("activity", "F4", "4.2","Asistent za ref. zakonodaje",      "Označevanje, barvno kodiranje",                 0.750, 19,  21),
    ("activity", "F4", "4.3","Asistent za ref. transkripta",     "Povezovanje izjav z dokumentacijo",             0.500, 19,  20),
    # F5
    ("phase",    "F5", "F5", "Frontend razvoj",                  "",                                               1.375, None, None),
    ("activity", "F5", "5.1","Ogrodje in navigacija",            "React, Azure AD, routing",                      0.250, 11,  11),
    ("activity", "F5", "5.2","Nadzorna plošča",                  "Kartice delovnih skupin",                       0.125, 12,  12),
    ("activity", "F5", "5.3","Delovna skupina — vmesniki",       "Dokumenti, člani, transkripti",                 0.125, 13,  13),
    ("activity", "F5", "5.4","Pregled transkriptov",             "Govorci, inline urejanje",                      0.250, 14,  14),
    ("activity", "F5", "5.5","Vmesniki asistentov",              "Split-pane, označevanje",                       0.250, 19,  19),
    ("activity", "F5", "5.6","Klepetalnik (Chatbot)",            "Floating widget",                               0.250, 20,  20),
    ("activity", "F5", "5.7","Administracija",                   "Uporabniki, zakonodaja, agenti",                0.125, 14,  14),
    # F6
    ("phase",    "F6", "F6", "Integracije",                      "",                                               0.625, None, None),
    ("activity", "F6", "6.1","Azure AD (SSO)",                   "Avtentikacija",                                 0.250, 11,  11),
    ("activity", "F6", "6.2","OneDrive sinhronizacija",          "Dvosmerna sinhronizacija",                      0.125, 13,  13),
    ("activity", "F6", "6.3","SRC transkripcija",                "Govor v besedilo",                              0.250, 13,  14),
    # F7
    ("phase",    "F7", "F7", "Skladnost in varnost",             "",                                               0.375, None, None),
    ("activity", "F7", "7.1","GDPR skladnost",                   "Klasifikacija, DPA",                            0.125, 3,    3),
    ("activity", "F7", "7.2","EU AI Act skladnost",              "Tveganja, razložljivost",                       0.250, 4,    5),
    # F8
    ("phase",    "F8", "F8", "Testiranje",                       "",                                               1.000, None, None),
    ("activity", "F8", "8.1","Funkcionalno testiranje",          "Testiranje modulov",                            0.250, 22,  22),
    ("activity", "F8", "8.2","AI testiranje",                    "Evalvacija kakovosti AI",                       0.250, 23,  23),
    ("activity", "F8", "8.4","UAT (uporabniško testiranje)",     "Testiranje z ATVP uporabniki",                  0.500, 24,  25),
    # F9
    ("phase",    "F9", "F9", "Uvedba in izobraževanje",          "",                                               1.000, None, None),
    ("activity", "F9", "9.1","Uvedba v produkcijo",              "Namestitev, migracija",                         0.250, 25,  25),
    ("activity", "F9", "9.3","Dokumentacija",                    "Tehnična dokumentacija",                        0.250, 23,  24),
    ("activity", "F9", "9.2","Izobraževanje uporabnikov",        "Delavnice za ATVP",                             0.500, 25,  26),
]

TOTAL_PM = 12.125
WEEKS = 26

# ─────────────────────────────────────────────
# Build workbook
# ─────────────────────────────────────────────
wb = Workbook()
ws = wb.active
ws.title = "Odin — Gantt"

# ── Column widths ──────────────────────────────
ws.column_dimensions["A"].width = 8
ws.column_dimensions["B"].width = 38
ws.column_dimensions["C"].width = 50
ws.column_dimensions["D"].width = 13
ws.column_dimensions["E"].width = 10
ws.column_dimensions["F"].width = 10
ws.column_dimensions["G"].width = 10
for w in range(WEEKS):
    col_letter = get_column_letter(8 + w)  # H=8
    ws.column_dimensions[col_letter].width = 3.5

# ── Row height default ─────────────────────────
ws.sheet_format.defaultRowHeight = 17

# ─────────────────────────────────────────────
# ROW 1 — Header
# ─────────────────────────────────────────────
HEADER_FILL  = fill("2D2D2D")
HEADER_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
WEEK_HDR_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=8)
center_align  = Alignment(horizontal="center", vertical="center", wrap_text=False)
left_align    = Alignment(horizontal="left",   vertical="center", wrap_text=False)
right_align   = Alignment(horizontal="right",  vertical="center")

headers_wbs = ["#", "Faza / Aktivnost", "Opis", "Obseg (PM)", "Začetek", "Konec", "Trajanje"]
for col_i, h in enumerate(headers_wbs, start=1):
    cell = ws.cell(row=1, column=col_i, value=h)
    cell.fill   = HEADER_FILL
    cell.font   = HEADER_FONT
    cell.alignment = center_align if col_i != 2 else left_align
    cell.border = thin_border()

for w in range(WEEKS):
    col_i = 8 + w
    cell = ws.cell(row=1, column=col_i, value=f"T{w+1}")
    cell.fill      = HEADER_FILL
    cell.font      = WEEK_HDR_FONT
    cell.alignment = center_align
    cell.border    = thin_border()

ws.row_dimensions[1].height = 20

# ─────────────────────────────────────────────
# DATA rows (starting at row 2)
# ─────────────────────────────────────────────
current_row = 2

for row_type, phase_key, number, name, desc, pm, t_start, t_end in WBS:
    p = PHASES[phase_key]
    r = current_row

    if row_type == "phase":
        # ── Phase row ─────────────────────────────
        bg = fill(p["phase_bg"])
        fnt_ph = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
        border = thin_border()

        ws.cell(r, 1, value=number).fill = bg
        ws.cell(r, 1).font = fnt_ph; ws.cell(r, 1).alignment = center_align; ws.cell(r, 1).border = border

        label = f"{number}  {name}"
        ws.cell(r, 2, value=label).fill = bg
        ws.cell(r, 2).font = fnt_ph; ws.cell(r, 2).alignment = left_align; ws.cell(r, 2).border = border

        ws.cell(r, 3, value="").fill = bg; ws.cell(r, 3).border = border

        pm_cell = ws.cell(r, 4, value=pm)
        pm_cell.fill = bg; pm_cell.font = fnt_ph
        pm_cell.alignment = center_align; pm_cell.border = border
        pm_cell.number_format = '0.000 "PM"'

        for col_i in [5, 6, 7]:
            ws.cell(r, col_i, value="").fill = bg; ws.cell(r, col_i).border = border

        # Gantt cells – dark phase BG
        for w in range(WEEKS):
            c = ws.cell(r, 8 + w, value="")
            c.fill = bg; c.border = thin_border()

        ws.row_dimensions[r].height = 18

    else:
        # ── Activity row ──────────────────────────
        bg      = fill(p["act_bg"])
        fnt_act = Font(name="Calibri", bold=False, color="2D2D2D", size=9)
        fnt_num = Font(name="Calibri", bold=False, color="555555", size=9)
        border  = thin_border()

        ws.cell(r, 1, value=number).fill = bg
        ws.cell(r, 1).font = fnt_num; ws.cell(r, 1).alignment = center_align; ws.cell(r, 1).border = border

        ws.cell(r, 2, value=f"   {name}").fill = bg
        ws.cell(r, 2).font = fnt_act; ws.cell(r, 2).alignment = left_align; ws.cell(r, 2).border = border

        ws.cell(r, 3, value=desc).fill = bg
        ws.cell(r, 3).font = Font(name="Calibri", color="555555", size=9, italic=True)
        ws.cell(r, 3).alignment = left_align; ws.cell(r, 3).border = border

        pm_cell = ws.cell(r, 4, value=pm)
        pm_cell.fill = bg; pm_cell.font = fnt_act
        pm_cell.alignment = center_align; pm_cell.border = border
        pm_cell.number_format = '0.000 "PM"'

        # Začetek / Konec / Trajanje
        t_start_str = f"T{t_start}" if t_start else ""
        t_end_str   = f"T{t_end}"   if t_end   else ""
        dur = (t_end - t_start + 1) if (t_start and t_end) else ""

        for col_i, val in [(5, t_start_str), (6, t_end_str), (7, dur)]:
            c = ws.cell(r, col_i, value=val)
            c.fill = bg; c.font = fnt_act
            c.alignment = center_align; c.border = border

        # Gantt bars
        is_11 = (number == "1.1")   # special lighter bar

        for w in range(WEEKS):
            week_num = w + 1
            c = ws.cell(r, 8 + w, value="")
            c.alignment = center_align
            c.border = thin_border()

            if t_start and t_end and t_start <= week_num <= t_end:
                if is_11:
                    bar_color = p.get("bar_light", "AED6F1")
                else:
                    bar_color = p["bar"]
                c.fill = fill(bar_color)
            else:
                c.fill = fill("F8F9FA")   # near-white background

        ws.row_dimensions[r].height = 16

    current_row += 1

# ─────────────────────────────────────────────
# Separator row (empty)
# ─────────────────────────────────────────────
sep_row = current_row
for col_i in range(1, 8 + WEEKS + 1):
    c = ws.cell(sep_row, col_i, value="")
    c.fill = fill("ECECEC")
    c.border = thin_border()
ws.row_dimensions[sep_row].height = 6
current_row += 1

# ─────────────────────────────────────────────
# TOTAL row
# ─────────────────────────────────────────────
tot_row = current_row
TOTAL_FILL = fill("2D2D2D")
TOTAL_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
border = thin_border()

ws.cell(tot_row, 1, value="").fill = TOTAL_FILL; ws.cell(tot_row, 1).border = border
ws.cell(tot_row, 2, value="SKUPAJ / TOTAL").fill = TOTAL_FILL
ws.cell(tot_row, 2).font = TOTAL_FONT; ws.cell(tot_row, 2).alignment = left_align; ws.cell(tot_row, 2).border = border
ws.cell(tot_row, 3, value="").fill = TOTAL_FILL; ws.cell(tot_row, 3).border = border

tot_pm = ws.cell(tot_row, 4, value=TOTAL_PM)
tot_pm.fill = TOTAL_FILL; tot_pm.font = TOTAL_FONT
tot_pm.alignment = center_align; tot_pm.border = border
tot_pm.number_format = '0.000 "PM"'

for col_i in [5, 6, 7]:
    c = ws.cell(tot_row, col_i, value="")
    c.fill = TOTAL_FILL; c.border = border

for w in range(WEEKS):
    c = ws.cell(tot_row, 8 + w, value="")
    c.fill = TOTAL_FILL; c.border = border

ws.row_dimensions[tot_row].height = 20

# ─────────────────────────────────────────────
# Freeze panes at H2 (row 1 header + cols A-G frozen)
# ─────────────────────────────────────────────
ws.freeze_panes = "H2"

# ─────────────────────────────────────────────
# Print setup: landscape, fit to 1 page wide
# ─────────────────────────────────────────────
ws.page_setup.orientation      = "landscape"
ws.page_setup.fitToWidth        = 1
ws.page_setup.fitToHeight       = 0
ws.page_setup.fitToPage         = True
ws.print_options.horizontalCentered = True
ws.page_margins.left   = 0.5
ws.page_margins.right  = 0.5
ws.page_margins.top    = 0.75
ws.page_margins.bottom = 0.75

# ─────────────────────────────────────────────
# Save
# ─────────────────────────────────────────────
import os, pathlib
out_path = pathlib.Path("/workspace/shared/odin-gantt.xlsx")
out_path.parent.mkdir(parents=True, exist_ok=True)
wb.save(str(out_path))
print(f"Saved → {out_path}")
print(f"Rows used: {current_row}")
